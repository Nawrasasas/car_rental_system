from django.contrib import admin
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.db.models import Q, Count, Sum
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse, Http404

from apps.accounting.models import Expense, EntryState
from apps.rentals.models import Rental, VehicleReplacement
from apps.vehicle_usage.models import VehicleUsage

from import_export.admin import ImportExportModelAdmin
from import_export import resources, fields
from import_export.widgets import ForeignKeyWidget
from apps.branches.models import Branch
from core.admin_site import custom_admin_site
from .models import Vehicle
from apps.attachments.inlines import AttachmentInline
from datetime import timedelta
from django.utils import timezone
from django.apps import apps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --- 1. Resource للاستيراد من الإكسل ---
from import_export import resources, fields
from import_export.widgets import (
    ForeignKeyWidget,
    BooleanWidget,
    DateWidget,
    DecimalWidget,
    IntegerWidget,
)

class VehicleResource(resources.ModelResource):
    # حاشية عربية: نبقي id أول عمود في الإكسل للعرض فقط، ويظل فارغًا عند الإضافة الجديدة
    id = fields.Field(column_name="id", attribute="id")

    # حاشية عربية: الفرع يُطابق على الاسم المكتوب في الإكسل مثل EIA / Downtown
    branch = fields.Field(
        column_name="branch",
        attribute="branch",
        widget=ForeignKeyWidget(Branch, "name"),
    )

    # حاشية عربية: حقول التاريخ
    registration_expiry = fields.Field(
        column_name="registration_expiry",
        attribute="registration_expiry",
        widget=DateWidget(format="%Y-%m-%d"),
    )
    annual_inspection_date = fields.Field(
        column_name="annual_inspection_date",
        attribute="annual_inspection_date",
        widget=DateWidget(format="%Y-%m-%d"),
    )
    insurance_expiry = fields.Field(
        column_name="insurance_expiry",
        attribute="insurance_expiry",
        widget=DateWidget(format="%Y-%m-%d"),
    )
    purchase_date = fields.Field(
        column_name="purchase_date",
        attribute="purchase_date",
        widget=DateWidget(format="%Y-%m-%d"),
    )
    last_service_date = fields.Field(
        column_name="last_service_date",
        attribute="last_service_date",
        widget=DateWidget(format="%Y-%m-%d"),
    )

    # حاشية عربية: الحقول الرقمية
    year = fields.Field(
        column_name="year", attribute="year", widget=IntegerWidget(), default=None)
    seats = fields.Field(column_name="seats", attribute="seats", widget=IntegerWidget())
    current_odometer = fields.Field(
        column_name="current_odometer",
        attribute="current_odometer",
        widget=IntegerWidget(),
    )
    last_service_odometer = fields.Field(
        column_name="last_service_odometer",
        attribute="last_service_odometer",
        widget=IntegerWidget(),
    )
    service_interval = fields.Field(
        column_name="service_interval",
        attribute="service_interval",
        widget=IntegerWidget(),
    )
    next_service_odometer = fields.Field(
        column_name="next_service_odometer",
        attribute="next_service_odometer",
        widget=IntegerWidget(),
    )
    key_count = fields.Field(
        column_name="key_count",
        attribute="key_count",
        widget=IntegerWidget(),
    )

    purchase_price = fields.Field(
        column_name="purchase_price",
        attribute="purchase_price",
        widget=DecimalWidget(),
    )
    daily_price = fields.Field(
        column_name="daily_price",
        attribute="daily_price",
        widget=DecimalWidget(),
    )
    weekly_price = fields.Field(
        column_name="weekly_price",
        attribute="weekly_price",
        widget=DecimalWidget(),
    )
    monthly_price = fields.Field(
        column_name="monthly_price",
        attribute="monthly_price",
        widget=DecimalWidget(),
    )
    deposit_amount = fields.Field(
        column_name="deposit_amount",
        attribute="deposit_amount",
        widget=DecimalWidget(),
    )
    extra_km_price = fields.Field(
        column_name="extra_km_price",
        attribute="extra_km_price",
        widget=DecimalWidget(),
    )

    # حاشية عربية: البوليان
    is_active = fields.Field(
        column_name="is_active",
        attribute="is_active",
        widget=BooleanWidget(),
    )

    class Meta:
        model = Vehicle

        # حاشية عربية: نبقي id أول عمود في الملف كما تريد
        fields = (
            "id",
            "branch",
            "plate_number",
            "brand",
            "model",
            "year",
            "vin_number",
            "engine_number",
            "color",
            "fuel_type",
            "transmission",
            "seats",
            "status",
            "registration_expiry",
            "annual_inspection_date",
            "insurance_company",
            "insurance_policy_number",
            "insurance_expiry",
            "ownership_type",
            "purchase_date",
            "purchase_price",
            "daily_price",
            "weekly_price",
            "monthly_price",
            "deposit_amount",
            "extra_km_price",
            "current_odometer",
            "last_service_odometer",
            "service_interval",
            "last_service_date",
            "next_service_odometer",
            "current_fuel_level",
            "key_count",
            "is_active",
            "notes",
        )

        export_order = fields

        # حاشية عربية: منع التكرار والمطابقة يكونان على رقم اللوحة لأنه unique
        import_id_fields = ("plate_number",)

        # حاشية عربية: نلغي منطق Skipped حتى تظهر الصفوف فعلًا في المعاينة
        skip_unchanged = False
        report_skipped = False

    def before_import(self, dataset, **kwargs):
        # حاشية عربية: تنظيف أسماء الأعمدة في الهيدر
        if dataset.headers:
            dataset.headers = [
                "" if h is None else str(h).strip() for h in dataset.headers
            ]
        return super().before_import(dataset, **kwargs)

    def before_import_row(self, row, **kwargs):
        # حاشية عربية: تنظيف النصوص وتحويل الفراغات إلى None
        for key, value in row.items():
            if isinstance(value, str):
                value = value.strip()
                row[key] = value if value != "" else None

        # حاشية عربية: نسمح بترك id فارغًا لأن النظام سيولده تلقائيًا
        if not row.get("id"):
            row["id"] = None

        # حاشية عربية: تجاهل الصفوف الفارغة بالكامل
        if not row.get("plate_number"):
            row["_skip_import"] = True
            return

        # حاشية عربية: الإلزامي الحقيقي حسب الموديل
        if not row.get("branch"):
            row["_skip_import"] = True
            return

        if not row.get("model"):
            row["_skip_import"] = True
            return

        if row.get("daily_price") in (None, ""):
            row["_skip_import"] = True
            return

        # حاشية عربية: توحيد رقم اللوحة لأنه مفتاح المطابقة
        row["plate_number"] = str(row["plate_number"]).strip()

        # حاشية عربية: defaults للحقول التي لا يجب أن تصل NULL
        if row.get("status") in (None, ""):
            row["status"] = "available"

        if row.get("ownership_type") in (None, ""):
            row["ownership_type"] = "company"

        if row.get("current_odometer") in (None, ""):
            row["current_odometer"] = 0

        if row.get("last_service_odometer") in (None, ""):
            row["last_service_odometer"] = 0

        if row.get("service_interval") in (None, ""):
            row["service_interval"] = 5000

        if row.get("key_count") in (None, ""):
            row["key_count"] = 1

        if row.get("is_active") in (None, ""):
            row["is_active"] = True

    def skip_row(self, instance, original, row, import_validation_errors=None):
        # حاشية عربية: فقط الصفوف الفارغة أو الناقصة إلزاميًا تُتخطى
        if row.get("_skip_import"):
            return True

        return super().skip_row(
            instance,
            original,
            row,
            import_validation_errors=import_validation_errors,
        )


class InsuranceRenewalFilter(admin.SimpleListFilter):
    # --- فلتر جانبي خاص بتنبيه التأمين ---
    title = "Insurance Status"
    parameter_name = "insurance_status"

    def lookups(self, request, model_admin):
        return (
            ("due_soon", "Insurance Due Soon"),
            ("expired", "Insurance Expired"),
        )

    def queryset(self, request, queryset):
        value = self.value()
        today = timezone.now().date()
        due_limit = today + timedelta(days=15)

        if value == "due_soon":
            # --- تأمين سينتهي خلال 15 يوم ولم ينتهِ بعد ---
            return queryset.filter(
                insurance_expiry__isnull=False,
                insurance_expiry__gte=today,
                insurance_expiry__lte=due_limit,
            )

        if value == "expired":
            # --- تأمين منتهي بالفعل ---
            return queryset.filter(
                insurance_expiry__isnull=False,
                insurance_expiry__lt=today,
            )

        return queryset


# --- 2. Vehicle Admin المطور (الداشبورد + تقسيم البطاقة إلى 3 أقسام) ---
@admin.register(Vehicle, site=custom_admin_site)
class VehicleAdmin(ImportExportModelAdmin):
    # --- تعريف الداشبورد الداخلي ---
    resource_class = VehicleResource
    change_list_template = "admin/vehicles/vehicle/change_list.html"
    # --- استخدام نظام المرفقات العام بدل مرفقات السيارة القديمة ---
    inlines = [AttachmentInline]
    # --- الأعمدة الظاهرة في قائمة السيارات ---
    list_display = (
        'plate_number',
        'brand',
        'model',
        'status_with_insurance',
        'current_rental_number',
        'current_rental_branch',
        'branch',
        'daily_price', 
    )

    # --- الفلاتر الجانبية ---
    list_filter = (
        "branch",
        "status",
        InsuranceRenewalFilter,
        "brand",
        "fuel_type",
        "transmission",
        "is_active",
    )

    # --- حقول البحث ---
    search_fields = (
        'plate_number',
        'brand',
        'model',
        'vin_number',
        'engine_number',
        'insurance_policy_number',
    )

    # --- الحقول القابلة للتعديل مباشرة من قائمة السيارات ---
    list_editable = (
        'daily_price',
        
    )

    # --- ترتيب صفحة البطاقة إلى 3 أقسام رئيسية ---
    fieldsets = (
        (
            "1) Vehicle Info",
            {
                "fields": (
                    ("plate_number", "branch", "status", "is_active"),
                    ("brand", "model", "year"),
                    ("vin_number", "engine_number"),
                    ("color", "fuel_type", "transmission", "seats"),
                )
            },
        ),
        ("4) Vehicle Record",
                 {"fields": (
                     ("vehicle_record_button"),
                )
            },
        ),
        (
            "2) Documents & Ownership",
            {
                "fields": (
                    ("registration_expiry", "annual_inspection_date"),
                    ("insurance_company", "insurance_policy_number"),
                    ("insurance_expiry", "ownership_type"),
                    ("purchase_date", "purchase_price"),
                )
            },
        ),
        (
            "3) Rental & Operations",
            {
                "fields": (
                    ("daily_price", "weekly_price", "monthly_price"),
                    ("deposit_amount", "extra_km_price"),
                    ("current_odometer", "current_fuel_level", "key_count"),
                    (
                        "last_service_odometer",
                        "service_interval",
                        "next_service_odometer",
                    ),
                    ("last_service_date",),
                    ("notes",),
                )
            },
        ),
    )

    class Media:
        # --- تحميل CSS العام للسيارات + CSS العام للمعرض ---
        css = {
            "all": (
                "css/admin_custom.css",
                "css/vehicle_admin_extras.css",
                "css/attachment_gallery_inline.css",
            )
        }

        # --- تحميل JS العام لمعاينة المرفقات قبل الحفظ ---
        js = ("js/attachment_gallery_inline.js",)

        # --- تحسين الاستعلامات لمنع كثرة الضرب على قاعدة البيانات ---
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.prefetch_related("rentals__branch")

    def current_rental_number(self, obj):
        # --- إظهار رقم العقد النشط الحالي للسيارة ---
        active_rental = obj.rentals.filter(status="active").order_by("-id").first()
        if active_rental:
            # نعرض رقم العقد الحقيقي بدل ID الداخلي لأن id يسبب تضارب مع رقم العقد الظاهر للمستخدم
            return active_rental.contract_number
        return "-"

    def status_with_insurance(self, obj):
        # --- تاريخ اليوم ---
        today = today = timezone.now().date()

        # --- حد التنبيه: قبل 15 يوم ---
        due_limit = today + timedelta(days=15)

        # --- نأخذ عرض الحالة الحالي ---
        # --- إذا كان عندك status_label قديم نستخدمه كما هو ---
        # --- نأخذ شارة الحالة الملونة الحالية بدل النص العادي ---
        if hasattr(self, "status_badge"):
            base_status = self.status_badge(obj)
        else:
            base_status = format_html("<span>{}</span>", obj.get_status_display())

        # --- إذا لا يوجد تاريخ تأمين فلا نضيف أي تنبيه ---
        if not obj.insurance_expiry:
            return base_status

        # --- إذا انتهى التأمين ---
        if obj.insurance_expiry < today:
            return format_html(
                '{} <span style="background:#dc2626; color:#fff; padding:2px 8px; '
                'border-radius:999px; font-size:11px; font-weight:600; margin-left:6px;">'
                'Insurance Expired'
                '</span>',
                base_status,
            )

        # --- إذا بقي 15 يوم أو أقل ---
        if today <= obj.insurance_expiry <= due_limit:
            return format_html(
                '{} <span style="background:#f59e0b; color:#111827; padding:2px 8px; '
                'border-radius:999px; font-size:11px; font-weight:600; margin-left:6px;">'
                'Insurance Due Soon'
                '</span>',
                base_status,
            )

        # --- الحالة الطبيعية بدون تنبيه ---
        return base_status

    status_with_insurance.short_description = "Status"

    current_rental_number.short_description = 'ACTIVE RENTAL'

    def current_rental_branch(self, obj):
        # --- إظهار فرع العقد النشط الحالي ---
        active_rental = obj.rentals.filter(status='active').order_by('-id').first()
        if active_rental and active_rental.branch:
            return active_rental.branch
        return '-'

    current_rental_branch.short_description = 'RENTAL BRANCH'

    def status_badge(self, obj):
        # --- ألوان الحالات الأساسية للسيارة ---
        colors = {
            "available": "#22c55e",
            "internal_use": "#7c3aed",
            "rented": "#ef4444",
            "maintenance": "#f59e0b",
            "service": "#06b6d4",
            "stolen": "#000000",
            "out_of_service": "#6b7280",
            "accident": "#b91c1c",
            "sold": "#1f2937",
        }

        # --- إنشاء شارة الحالة الأساسية مثل Available أو Rented ---
        base_color = colors.get(obj.status, "#6c757d")
        badges = [
            format_html(
                '<span style="background:{}; color:white; padding:5px 12px; '
                'border-radius:15px; font-weight:bold; font-size:11px; '
                'display:inline-block; margin-right:6px;">{}</span>',
                base_color,
                obj.get_status_display(),
            )
        ]

        # --- جلب آخر عقد نشط للسيارة إن وجد ---
        active_rental = obj.rentals.filter(status='active').order_by('-id').first()

        # --- إذا السيارة مؤجرة والعقد متأخر نضيف شارة Overdue بجانب الحالة الأساسية ---
        if active_rental and active_rental.is_overdue:
            badges.append(
                mark_safe(
                    '<span style="background:#dc2626; color:white; padding:5px 12px; '
                    'border-radius:15px; font-weight:bold; font-size:11px; '
                    'display:inline-block; margin-right:6px;">Overdue</span>'
                )
            )

        # --- إذا السيارة متاحة لكنها تحتاج صيانة نضيف شارة Service Due بجانب Available ---
        if obj.status == "available" and obj.needs_service:
            badges.append(
                mark_safe(
                    '<span style="background:#f97316; color:white; padding:5px 12px; '
                    'border-radius:15px; font-weight:bold; font-size:11px; '
                    'display:inline-block;">Service Due</span>'
                )
            )

        # --- دمج كل الشارات في خلية واحدة ---
        return mark_safe(" ".join(str(badge) for badge in badges))

    status_badge.short_description = "Status"

    def vehicle_record_button(self, obj):
        # --- زر فتح سجل السيارة الموحد من داخل بطاقة السيارة ---
        if not obj or not obj.pk:
            return "Save the vehicle first to open the vehicle record."

        url = reverse(
            f"{self.admin_site.name}:vehicles_vehicle_record",
            args=[obj.pk],
        )

        return format_html(
            '<a href="{}" '
            'style="background:#0f766e; color:white; padding:10px 16px; '
            'border-radius:8px; text-decoration:none; font-weight:600; display:inline-block;">'
            'Open Vehicle Record'
            '</a>',
            url,
        )

    vehicle_record_button.short_description = "Vehicle Record"

    def _format_vehicle_record_datetime(self, value):
        # --- توحيد عرض التاريخ/الوقت داخل سجل السيارة ---
        if not value:
            return "-"

        if timezone.is_aware(value):
            value = timezone.localtime(value)

        return value.strftime("%Y-%m-%d %H:%M")

    def _vehicle_record_sort_key(self, value):
        # --- مفتاح ترتيب نصي آمن لتجميع المصادر المختلفة في جدول واحد ---
        if not value:
            return ""

        if timezone.is_aware(value):
            value = timezone.localtime(value)

        return value.strftime("%Y%m%d%H%M%S")

    def _display_record_user(self, user):
        # --- عرض اسم المستخدم بشكل لطيف داخل التفاصيل ---
        if not user:
            return "-"
        return user.get_full_name() or user.username

    def _build_vehicle_record_rows(self, vehicle):
        # --- نجمع كل مصادر السجل هنا: Rental + VehicleUsage + VehicleReplacement ---
        rows = []

        # ======================================================
        # 1) Rental rows
        # ======================================================
        rentals = (
            vehicle.rentals
            .select_related("customer", "branch", "created_by")
            .all()
        )

        for rental in rentals:
            details = []

            if rental.created_by:
                details.append(
                    f"Created by: {self._display_record_user(rental.created_by)}"
                )

            if getattr(rental, "online_reference", None):
                details.append(f"Online Ref: {rental.online_reference}")

            if getattr(rental, "contract_notes", None):
                details.append(rental.contract_notes)

            reference = rental.contract_number or f"Rental #{rental.pk}"

            rows.append(
                {
                    "source": "rental",
                    "record_type": "Rental",
                    "reference": reference,
                    "related_contract": reference,
                    "role": "Primary Vehicle",
                    "from_branch": rental.branch.name if rental.branch else "-",
                    "to_branch": rental.branch.name if rental.branch else "-",
                    "time_out": self._format_vehicle_record_datetime(rental.start_date),
                    "time_in": self._format_vehicle_record_datetime(
                        rental.actual_return_date or rental.end_date
                    ),
                    "km_out": rental.pickup_odometer if rental.pickup_odometer is not None else "-",
                    "km_in": rental.return_odometer if rental.return_odometer is not None else "-",
                    "status": rental.display_status,
                    "person": str(rental.customer) if rental.customer else "-",
                    "details": " | ".join(details) if details else "-",
                    "open_url": reverse(
                        f"{self.admin_site.name}:rentals_rental_change",
                        args=[rental.pk],
                    ),
                    "open_label": "Open Rental",
                    "sort_key": self._vehicle_record_sort_key(
                        rental.actual_return_date or rental.start_date
                    ),
                }
            )

        # ======================================================
        # 2) VehicleUsage rows
        # ======================================================
        usages = (
            vehicle.vehicle_usages
            .select_related("source_branch", "destination_branch", "created_by")
            .all()
        )

        for usage in usages:
            usage_status = usage.get_status_display()

            if (
                usage.status == VehicleUsage.STATUS_ACTIVE
                and usage.expected_return_datetime
                and usage.expected_return_datetime < timezone.now()
            ):
                usage_status = f"{usage_status} / Overdue"

            details = []

            if usage.handover_by:
                details.append(f"Handed by: {usage.handover_by}")

            if usage.received_by:
                details.append(f"Received by: {usage.received_by}")

            if usage.created_by:
                details.append(
                    f"Created by: {self._display_record_user(usage.created_by)}"
                )

            if usage.notes:
                details.append(usage.notes)

            rows.append(
                {
                    "source": "usage",
                    "record_type": "Vehicle Usage",
                    "reference": usage.usage_no or f"Usage #{usage.pk}",
                    "related_contract": "-",
                    "role": usage.get_purpose_display(),
                    "from_branch": usage.source_branch.name if usage.source_branch else "-",
                    "to_branch": usage.destination_branch.name if usage.destination_branch else "-",
                    "time_out": self._format_vehicle_record_datetime(usage.start_datetime),
                    "time_in": self._format_vehicle_record_datetime(
                        usage.actual_return_datetime or usage.expected_return_datetime
                    ),
                    "km_out": usage.pickup_odometer if usage.pickup_odometer is not None else "-",
                    "km_in": usage.return_odometer if usage.return_odometer is not None else "-",
                    "status": usage_status,
                    "person": usage.employee_name or "-",
                    "details": " | ".join(details) if details else "-",
                    "open_url": reverse(
                        f"{self.admin_site.name}:vehicle_usage_vehicleusage_change",
                        args=[usage.pk],
                    ),
                    "open_label": "Open Usage",
                    "sort_key": self._vehicle_record_sort_key(
                        usage.actual_return_datetime or usage.start_datetime
                    ),
                }
            )

        # ======================================================
        # 3) VehicleReplacement rows
        # ======================================================
        replacements = (
            VehicleReplacement.objects.filter(
                Q(original_vehicle_id=vehicle.pk) | Q(replacement_vehicle_id=vehicle.pk)
            )
            .select_related(
                "rental",
                "rental__customer",
                "rental__branch",
                "original_vehicle",
                "replacement_vehicle",
                "started_by",
                "ended_by",
            )
            .all()
        )

        for replacement in replacements:
            contract_ref = (
                replacement.rental.contract_number
                or f"Rental #{replacement.rental_id}"
            )

            is_original_vehicle = replacement.original_vehicle_id == vehicle.pk

            if is_original_vehicle:
                role = "Original Vehicle"
                from_display = replacement.rental.branch.name if replacement.rental.branch else "-"
                to_display = f"Temporarily Replaced by {replacement.replacement_vehicle.plate_number}"
                main_details = (
                    f"Original vehicle replaced by "
                    f"{replacement.replacement_vehicle.plate_number}. "
                    f"Reason: {replacement.reason}"
                )
            else:
                role = "Replacement Vehicle"
                from_display = "Available Fleet"
                to_display = f"Customer / {contract_ref}"
                main_details = (
                    f"Sent as replacement for original vehicle "
                    f"{replacement.original_vehicle.plate_number}. "
                    f"Reason: {replacement.reason}"
                )

            details = [main_details]

            if replacement.started_by:
                details.append(
                    f"Started by: {self._display_record_user(replacement.started_by)}"
                )

            if replacement.ended_by:
                details.append(
                    f"Ended by: {self._display_record_user(replacement.ended_by)}"
                )

            if replacement.notes:
                details.append(replacement.notes)

            rows.append(
                {
                    "source": "replacement",
                    "record_type": "Vehicle Replacement",
                    "reference": f"VR-{replacement.pk}",
                    "related_contract": contract_ref,
                    "role": role,
                    "from_branch": from_display,
                    "to_branch": to_display,
                    "time_out": self._format_vehicle_record_datetime(replacement.started_at),
                    "time_in": self._format_vehicle_record_datetime(replacement.ended_at),
                    "km_out": "-",
                    "km_in": "-",
                    "status": replacement.get_status_display(),
                    "person": str(replacement.rental.customer) if replacement.rental.customer else "-",
                    "details": " | ".join(details),
                    "open_url": reverse(
                        f"{self.admin_site.name}:rentals_rental_change",
                        args=[replacement.rental_id],
                    ),
                    "open_label": "Open Rental",
                    "sort_key": self._vehicle_record_sort_key(
                        replacement.ended_at or replacement.started_at
                    ),
                }
            )

        # --- ترتيب نهائي تنازلي بحيث يظهر الأحدث أولًا ---
        rows.sort(key=lambda row: row["sort_key"], reverse=True)
        return rows

    def vehicle_record_view(self, request, object_id):
        # --- صفحة مستقلة داخل الأدمن لعرض سجل السيارة الموحد ---
        vehicle = self.get_object(request, object_id)

        if vehicle is None:
            raise Http404("Vehicle not found.")

        if not (
            self.has_view_permission(request, vehicle)
            or self.has_change_permission(request, vehicle)
        ):
            raise PermissionDenied

        rows = self._build_vehicle_record_rows(vehicle)

        context = dict(
            self.admin_site.each_context(request),
            title=f"Vehicle Record - {vehicle.plate_number}",
            opts=self.model._meta,
            original=vehicle,
            vehicle=vehicle,
            rows=rows,
            record_summary={
                "all": len(rows),
                "rentals": sum(1 for row in rows if row["source"] == "rental"),
                "usages": sum(1 for row in rows if row["source"] == "usage"),
                "replacements": sum(1 for row in rows if row["source"] == "replacement"),
            },
            admin_index_url=reverse(f"{self.admin_site.name}:index"),
            app_list_url=reverse(
                f"{self.admin_site.name}:app_list",
                kwargs={"app_label": self.model._meta.app_label},
            ),
            vehicle_changelist_url=reverse(
                f"{self.admin_site.name}:vehicles_vehicle_changelist"
            ),
            vehicle_change_url=reverse(
                f"{self.admin_site.name}:vehicles_vehicle_change",
                args=[vehicle.pk],
            ),
        )

        return TemplateResponse(
            request,
            "admin/vehicles/vehicle/vehicle_record.html",
            context,
        )

    def get_urls(self):
        # --- إضافة رابط سجل السيارة الموحد مع الإبقاء على الروابط الحالية ---
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/vehicle-record/",
                self.admin_site.admin_view(self.vehicle_record_view),
                name="vehicles_vehicle_record",
            ),
            path(
                "vehicle-expense-ranking/",
                self.admin_site.admin_view(self.vehicle_expense_ranking_view),
                name="vehicles_vehicle_expense_ranking",
            ),
            path(
                "vehicle-expense-ranking/export-excel/",
                self.admin_site.admin_view(self.export_vehicle_expense_ranking_excel),
                name="vehicles_vehicle_expense_ranking_export_excel",
            ),
        ]
        return custom_urls + urls

    def get_vehicle_expense_ranking_queryset(self, from_date=None, to_date=None):
        # --- هذا هو مصدر الحقيقة الوحيد للتقرير والتصدير حتى لا تختلف نتائج الشاشة عن الإكسل ---
        queryset = Expense.objects.filter(
            vehicle__isnull=False,
            state=EntryState.POSTED,
        )

        # --- تطبيق نفس الفلاتر الزمنية على التقرير والتصدير ---
        if from_date:
            queryset = queryset.filter(expense_date__gte=from_date)

        if to_date:
            queryset = queryset.filter(expense_date__lte=to_date)

        # --- تجميع المصروفات حسب السيارة وترتيبها من الأعلى للأقل ---
        return (
            queryset.values(
                "vehicle_id",
                "vehicle__plate_number",
                "vehicle__brand",
                "vehicle__model",
                "vehicle__branch__name",
            )
            .annotate(
                total_expense=Sum("amount"),
                expense_count=Count("id"),
            )
            .order_by("-total_expense", "vehicle__plate_number")
        )

    def vehicle_expense_ranking_view(self, request):
        # --- قراءة فترة التقرير من الرابط ---
        from_date = request.GET.get("from_date")
        to_date = request.GET.get("to_date")

        # --- جلب نفس بيانات التقرير من الدالة المشتركة ---
        ranking = self.get_vehicle_expense_ranking_queryset(
            from_date=from_date,
            to_date=to_date,
        )

        # --- تجهيز رابط التصدير مع نفس فلاتر الصفحة الحالية ---
        export_excel_url = reverse(
            "admin:vehicles_vehicle_expense_ranking_export_excel"
        )
        current_params = request.GET.copy()
        if current_params:
            export_excel_url = f"{export_excel_url}?{current_params.urlencode()}"

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Vehicle Expense Ranking",
            "ranking": ranking,
            "from_date": from_date,
            "to_date": to_date,
            "back_url": reverse("admin:vehicles_vehicle_changelist"),
            "export_excel_url": export_excel_url,
        }

        return TemplateResponse(
            request,
            "admin/accounting/expense/vehicle_expense_ranking.html",
            context,
        )

    def export_vehicle_expense_ranking_excel(self, request):
        # --- قراءة نفس فلاتر التقرير من الرابط ---
        from_date = request.GET.get("from_date")
        to_date = request.GET.get("to_date")

        # --- جلب نفس النتائج المعروضة في الشاشة ---
        ranking = self.get_vehicle_expense_ranking_queryset(
            from_date=from_date,
            to_date=to_date,
        )

        # --- إنشاء ملف Excel ---
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Vehicle Expense Ranking"

        # --- عنوان التقرير ---
        worksheet.merge_cells("A1:G1")
        worksheet["A1"] = "Vehicle Expense Ranking"
        worksheet["A1"].font = Font(size=14, bold=True)
        worksheet["A1"].alignment = Alignment(horizontal="center")

        # --- معلومات الفترة المحددة ---
        worksheet.merge_cells("A2:G2")
        worksheet["A2"] = f"Period: {from_date or 'Beginning'} -> {to_date or 'Today'}"
        worksheet["A2"].font = Font(bold=True)
        worksheet["A2"].alignment = Alignment(horizontal="center")

        # --- رؤوس الأعمدة ---
        headers = [
            "#",
            "Vehicle No",
            "Brand",
            "Model",
            "Branch",
            "Expense Count",
            "Total Expense",
        ]

        header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")

        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=4, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # --- تعبئة الصفوف ---
        for row_num, row in enumerate(ranking, start=5):
            worksheet.cell(row=row_num, column=1, value=row_num - 4)
            worksheet.cell(
                row=row_num, column=2, value=row.get("vehicle__plate_number") or "-"
            )
            worksheet.cell(
                row=row_num, column=3, value=row.get("vehicle__brand") or "-"
            )
            worksheet.cell(
                row=row_num, column=4, value=row.get("vehicle__model") or "-"
            )
            worksheet.cell(
                row=row_num, column=5, value=row.get("vehicle__branch__name") or "-"
            )
            worksheet.cell(row=row_num, column=6, value=row.get("expense_count") or 0)

            total_expense_value = float(row.get("total_expense") or 0)
            total_cell = worksheet.cell(
                row=row_num, column=7, value=total_expense_value
            )
            total_cell.number_format = "#,##0.00"

        # --- تثبيت صف العناوين ---
        worksheet.freeze_panes = "A5"

        # --- ضبط عرض الأعمدة ---
        worksheet.column_dimensions["A"].width = 8
        worksheet.column_dimensions["B"].width = 18
        worksheet.column_dimensions["C"].width = 18
        worksheet.column_dimensions["D"].width = 18
        worksheet.column_dimensions["E"].width = 18
        worksheet.column_dimensions["F"].width = 16
        worksheet.column_dimensions["G"].width = 18

        # --- اسم الملف المصدّر ---
        file_from = from_date or "beginning"
        file_to = to_date or "today"
        filename = f"vehicle_expense_ranking_{file_from}_to_{file_to}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        workbook.save(response)
        return response

    def changelist_view(self, request, extra_context=None):
        # --- تخصيص صفحة قائمة السيارات مع إحصائيات أعلى الصفحة ---
        extra_context = extra_context or {}
        extra_context["vehicle_expense_ranking_url"] = reverse(
            "admin:accounting_expense_vehicle_expense_ranking"
        )

        # --- جلب الكويري الحالي بعد تطبيق أي فلتر من المستخدم ---
        cl = self.get_changelist_instance(request)

        # --- نستخدم root_queryset مثل صفحة العقود ---
        # --- حتى تبقى أرقام كل البطاقات صحيحة عند الضغط على أي حالة ---
        base_qs = cl.root_queryset

        # --- نحسب العدادات من الأساس العام للقائمة ---
        # --- وليس من queryset المفلتر بالحالة الحالية ---
        stats = base_qs.aggregate(
            total=Count("id"),
            available=Count("id", filter=Q(status="available")),
            internal_use=Count("id", filter=Q(status="internal_use")),
            rented=Count("id", filter=Q(status="rented")),
            maintenance=Count(
                "id", filter=Q(status="maintenance")
            ),  # حاشية عربية: عدّ مستقل لحالة maintenance
            service=Count(
                "id", filter=Q(status="service")
            ),  # حاشية عربية: عدّ مستقل لحالة service
            out_of_service=Count("id", filter=Q(status="out_of_service")),
            accident=Count("id", filter=Q(status="accident")),
            stolen=Count("id", filter=Q(status="stolen")),
        )

        # --- نحتفظ ببقية الفلاتر الحالية ونزيل رقم الصفحة ---
        current_params = request.GET.copy()
        current_params.pop("p", None)

        # --- دالة مساعدة لإنشاء روابط الفلترة السريعة ---
        def get_filter_url(status_val):
            params = current_params.copy()
            if status_val:
                params['status__exact'] = status_val
            elif 'status__exact' in params:
                del params['status__exact']
            return f'?{params.urlencode()}'

        # --- HTML مبسط لبطاقات الإحصائيات أعلى الصفحة ---
        extra_context["dashboard_html"] = mark_safe(
            f"""
            <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 15px; margin-bottom: 25px;">

                <a href="{get_filter_url(None)}" style="text-decoration: none;">
                    <div style="background: white; padding: 15px; border-radius: 12px; text-align: center; border-bottom: 4px solid #6366f1; box-shadow: 0 4px 6px rgba(0,0,0,0.07); transition: transform 0.2s;" onmouseover="this.style.transform='translateY(-3px)'" onmouseout="this.style.transform='translateY(0)'">
                        <div style="font-size: 26px; font-weight: bold; color: #1e1b4b;">{stats['total']}</div>
                        <div style="font-size: 12px; color: #6366f1; font-weight: bold; text-transform: uppercase; margin-top: 5px;">Total Fleet</div>
                    </div>
                </a>

                <a href="{get_filter_url('available')}" style="text-decoration: none;">
                    <div style="background: #f0fdf4; padding: 15px; border-radius: 12px; text-align: center; border-bottom: 4px solid #22c55e; box-shadow: 0 4px 6px rgba(0,0,0,0.07); transition: transform 0.2s;" onmouseover="this.style.transform='translateY(-3px)'" onmouseout="this.style.transform='translateY(0)'">
                        <div style="font-size: 26px; font-weight: bold; color: #14532d;">{stats['available']}</div>
                        <div style="font-size: 12px; color: #16a34a; font-weight: bold; text-transform: uppercase; margin-top: 5px;">Available</div>
                    </div>
                </a>

                                <a href="{get_filter_url('internal_use')}" style="text-decoration: none;">
                    <div style="background: #f5f3ff; padding: 15px; border-radius: 12px; text-align: center; border-bottom: 4px solid #7c3aed; box-shadow: 0 4px 6px rgba(0,0,0,0.07); transition: transform 0.2s;" onmouseover="this.style.transform='translateY(-3px)'" onmouseout="this.style.transform='translateY(0)'">
                        <div style="font-size: 26px; font-weight: bold; color: #4c1d95;">{stats['internal_use']}</div>
                        <div style="font-size: 12px; color: #7c3aed; font-weight: bold; text-transform: uppercase; margin-top: 5px;">Internal Use</div>
                    </div>
                </a>
                
                <a href="{get_filter_url('rented')}" style="text-decoration: none;">
                    <div style="background: #fef2f2; padding: 15px; border-radius: 12px; text-align: center; border-bottom: 4px solid #ef4444; box-shadow: 0 4px 6px rgba(0,0,0,0.07); transition: transform 0.2s;" onmouseover="this.style.transform='translateY(-3px)'" onmouseout="this.style.transform='translateY(0)'">
                        <div style="font-size: 26px; font-weight: bold; color: #7f1d1d;">{stats['rented']}</div>
                        <div style="font-size: 12px; color: #dc2626; font-weight: bold; text-transform: uppercase; margin-top: 5px;">Rented</div>
                    </div>
                </a>

                <a href="{get_filter_url('maintenance')}" style="text-decoration: none;">
                    <div style="background: #fffbeb; padding: 15px; border-radius: 12px; text-align: center; border-bottom: 4px solid #f59e0b; box-shadow: 0 4px 6px rgba(0,0,0,0.07); transition: transform 0.2s;" onmouseover="this.style.transform='translateY(-3px)'" onmouseout="this.style.transform='translateY(0)'">
                        <div style="font-size: 26px; font-weight: bold; color: #78350f;">{stats['maintenance']}</div>  <!-- حاشية: هنا يجب عرض عدد maintenance فقط لأن الرابط يفلتر maintenance فقط -->
                        <div style="font-size: 12px; color: #d97706; font-weight: bold; text-transform: uppercase; margin-top: 5px;">Maintenance</div>  <!-- حاشية: غيّرنا النص حتى يطابق الفلتر والعدد المعروض -->
                    </div>
                </a>

                <a href="{get_filter_url('service')}" style="text-decoration: none;">  <!-- حاشية: أضفنا بطاقة مستقلة لـ service لأن جمعها مع maintenance كان يسبب تضليلًا -->
                    <div style="background: #ecfeff; padding: 15px; border-radius: 12px; text-align: center; border-bottom: 4px solid #06b6d4; box-shadow: 0 4px 6px rgba(0,0,0,0.07); transition: transform 0.2s;" onmouseover="this.style.transform='translateY(-3px)'" onmouseout="this.style.transform='translateY(0)'">
                        <div style="font-size: 26px; font-weight: bold; color: #155e75;">{stats['service']}</div>  <!-- حاشية: هنا نعرض عدد service فقط -->
                        <div style="font-size: 12px; color: #0891b2; font-weight: bold; text-transform: uppercase; margin-top: 5px;">Service</div>  <!-- حاشية: النص الآن مطابق تمامًا للفلتر -->
                    </div>
                </a>

            </div>
        """
        )
        extra_context["vehicle_expense_ranking_url"] = reverse(
        "admin:vehicles_vehicle_expense_ranking"
        )
        return super().changelist_view(request, extra_context=extra_context)

    def get_readonly_fields(self, request, obj=None):
        readonly = list(super().get_readonly_fields(request, obj))

        # --- زر السجل الموحد يظهر كحقل قراءة فقط دائمًا ---
        readonly.append("vehicle_record_button")

        # --- إذا السيارة لديها عقد نشط نمنع تعديل الحالة ---
        if obj and obj.rentals.filter(status="active").exists():
            readonly.append("status")

        return tuple(dict.fromkeys(readonly))


def get_search_results(self, request, queryset, search_term):
    queryset, use_distinct = super().get_search_results(request, queryset, search_term)

    app_label = request.GET.get("app_label")
    model_name = request.GET.get("model_name")
    field_name = request.GET.get("field_name")
    object_id = request.resolver_match.kwargs.get("object_id")

    # --- نعتمد فقط على حالة السيارة الحالية ---
    if app_label == "rentals" and model_name == "rental" and field_name == "vehicle":
        if object_id:
            try:
                rental = Rental.objects.only("vehicle_id").get(pk=object_id)
                queryset = queryset.filter(
                    Q(status="available") | Q(pk=rental.vehicle_id)
                )
            except Rental.DoesNotExist:
                queryset = queryset.filter(status="available")
        else:
            queryset = queryset.filter(status="available")

    # --- نفس المنطق في Vehicle Usage ---
    elif (
        app_label == "vehicle_usage"
        and model_name == "vehicleusage"
        and field_name == "vehicle"
    ):
        VehicleUsage = apps.get_model("vehicle_usage", "VehicleUsage")

        if object_id:
            try:
                usage = VehicleUsage.objects.only("vehicle_id").get(pk=object_id)
                queryset = queryset.filter(
                    Q(status="available") | Q(pk=usage.vehicle_id)
                )
            except VehicleUsage.DoesNotExist:
                queryset = queryset.filter(status="available")
        else:
            queryset = queryset.filter(status="available")

    return queryset, use_distinct
