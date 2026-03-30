from django.contrib import admin
from django.urls import path
from django.template.response import TemplateResponse
from django.utils import timezone
from django.http import HttpResponse
from openpyxl import Workbook

# استدعاء موقع الأدمن المخصص الخاص بك
from core.admin_site import custom_admin_site

# استدعاء النماذج الوهمية والخدمات
from .models import GeneralLedger, IncomeStatement, SalesReport
from .services import get_income_statement, get_general_ledger, get_sales_report

# إذا كان لديك دالة مخصصة للمبيعات، قم باستدعائها هنا
from django import forms


class SalesReportFilterForm(forms.Form):
    start_date = forms.DateField(
        required=False,
        label="Start Date",
        widget=forms.DateInput(attrs={"type": "date"}),
    )
    end_date = forms.DateField(
        required=False,
        label="End Date",
        widget=forms.DateInput(attrs={"type": "date"}),
    )
    status = forms.ChoiceField(
        required=False,
        label="Status",
        choices=[
            ("", "All"),
            ("active", "Active"),
            ("completed", "Completed"),
        ],
    )


@admin.register(SalesReport, site=custom_admin_site)
class SalesReportAdmin(admin.ModelAdmin):

    def get_urls(self):
        urls = super().get_urls()
        # نستخدم نفس الطريقة هنا لاعتراض الرابط والبقاء داخل الـ Admin
        custom_urls = [
            path(
                "",
                self.admin_site.admin_view(self.sales_dashboard),
                name="reports_salesreport_changelist",
            ),
        ]
        return custom_urls + urls

    def export_sales_excel(self, rentals, total_sales):
        # --- إنشاء ملف Excel جديد ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # --- العناوين الرئيسية ---
        ws.append([
            "Rental ID",
            "Contract Number",
            "Vehicle",
            "Customer",
            "Start Date",
            "Status",
            "Net Total",
        ])

        # --- تعبئة الصفوف من نتائج التقرير الحالية ---
        for rental in rentals:
            ws.append([
                rental.id,
                getattr(rental, "contract_number", "") or "",
                getattr(rental.vehicle, "plate_number", "") if getattr(rental, "vehicle", None) else "",
                str(rental.customer) if getattr(rental, "customer", None) else "",
                rental.start_date.strftime("%Y-%m-%d") if rental.start_date else "",
                rental.status,
                float(rental.net_total or 0),
            ])

        # --- سطر إجمالي المبيعات في آخر الملف ---
        ws.append([])
        ws.append(["", "", "", "", "", "Total Sales", float(total_sales or 0)])

        # --- تجهيز الاستجابة للتحميل ---
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="sales_report.xlsx"'

        wb.save(response)
        return response

    def sales_dashboard(self, request):
        # 1. جلب السياق الافتراضي للحفاظ على شكل القائمة الجانبية (Sidebar)
        context = dict(self.admin_site.each_context(request))
        context["title"] = "Sales Report"
        today = timezone.now().date()
        start_of_month = today.replace(day=1)

        raw_data = request.GET.copy()
        if not raw_data:
            raw_data = {
                "start_date": start_of_month.isoformat(),
                "end_date": today.isoformat(),
                "status": "",
            }

        form = SalesReportFilterForm(raw_data)
        context["form"] = form
        context["rentals"] = None
        context["total_sales"] = 0

        if form.is_valid():
            sales_data = get_sales_report(
                start_date=form.cleaned_data.get("start_date"),
                end_date=form.cleaned_data.get("end_date"),
                status=form.cleaned_data.get("status") or None,
            )
            context["rentals"] = sales_data["rentals"]
            context["total_sales"] = sales_data["total_sales"]

            # --- إذا ضغط المستخدم زر التصدير إلى Excel نرجع الملف مباشرة ---
            if request.GET.get("export") == "excel":
                return self.export_sales_excel(
                    rentals=context["rentals"],
                    total_sales=context["total_sales"],
                )

        # توجيه المستخدم لصفحة HTML مخصصة للمبيعات
        return TemplateResponse(request, "admin/reports/sales_report.html", context)


@admin.register(IncomeStatement, site=custom_admin_site)
class IncomeStatementAdmin(admin.ModelAdmin):

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "",
                self.admin_site.admin_view(self.income_dashboard),
                name="reports_incomestatement_changelist",
            ),
        ]
        return custom_urls + urls
    
    def export_income_statement_excel(self, income_data):
        wb = Workbook()
        ws = wb.active
        ws.title = "Income Statement"

        ws.append(["Income Statement"])
        ws.append(["Period Start", str(income_data["period_start"])])
        ws.append(["Period End", str(income_data["period_end"])])
        ws.append([])

        ws.append(["REVENUES"])
        ws.append(["Code", "Account Name", "Balance"])

        for rev in income_data["revenues"]:
            ws.append([
                rev["code"],
                rev["name"],
                float(rev["balance"] or 0),
            ])

        ws.append(["", "Total Revenues", float(income_data["total_revenue"] or 0)])
        ws.append([])

        ws.append(["EXPENSES"])
        ws.append(["Code", "Account Name", "Balance"])

        for exp in income_data["expenses"]:
            ws.append([
                exp["code"],
                exp["name"],
                float(exp["balance"] or 0),
            ])

        ws.append(["", "Total Expenses", float(income_data["total_expense"] or 0)])
        ws.append([])
        ws.append(["", "Net Income", float(income_data["net_income"] or 0)])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="income_statement.xlsx"'
        wb.save(response)
        return response

    def income_dashboard(self, request):
        context = dict(self.admin_site.each_context(request))
        context["title"] = "Income Statement"

        today = timezone.now().date()
        start_of_month = today.replace(day=1)

        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        # إذا لم يُدخل المستخدم شيئًا، نستخدم افتراضيًا أول الشهر إلى اليوم
        if not start_date:
            start_date = start_of_month.isoformat()

        if not end_date:
            end_date = today.isoformat()

        income_data = get_income_statement(
            start_date=start_date,
            end_date=end_date,
        )

        context["income_data"] = income_data
        if request.GET.get("export") == "excel":
            return self.export_income_statement_excel(income_data)
        return TemplateResponse(request, "admin/reports/income_statement.html", context)


@admin.register(GeneralLedger, site=custom_admin_site)
class GeneralLedgerAdmin(admin.ModelAdmin):

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "",
                self.admin_site.admin_view(self.ledger_dashboard),
                name="reports_generalledger_changelist",
            ),
        ]
        return custom_urls + urls


    def export_general_ledger_excel(self, ledger_data):
        wb = Workbook()
        ws = wb.active
        ws.title = "General Ledger"

        ws.append(["General Ledger"])
        ws.append(["Account Name", ledger_data["account_name"]])
        ws.append(["Account Code", ledger_data["account_code"]])
        ws.append(["Opening Balance", float(ledger_data["opening_balance"] or 0)])
        ws.append([])

        ws.append([
            "Date",
            "Entry No.",
            "Description",
            "Source",
            "Debit",
            "Credit",
            "Running Balance",
        ])

        for line in ledger_data["transactions"]:
            ws.append([
                str(line["date"]) if line.get("date") else "",
                line.get("entry_number", ""),
                line.get("description", ""),
                line.get("source_label", ""),
                float(line.get("debit", 0) or 0),
                float(line.get("credit", 0) or 0),
                float(line.get("balance", 0) or 0),
            ])

        ws.append([])
        ws.append(["", "", "", "Closing Balance", "", "", float(ledger_data["closing_balance"] or 0)])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="general_ledger.xlsx"'
        wb.save(response)
        return response

    def ledger_dashboard(self, request):
        context = dict(self.admin_site.each_context(request))
        context["title"] = "General Ledger"

        # Read GET parameters from the URL
        account_code = request.GET.get("account_code")
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        if account_code:
            try:
                # Call our service engine with user inputs
                ledger_data = get_general_ledger(
                    account_code=account_code,
                    start_date=start_date if start_date else None,
                    end_date=end_date if end_date else None,
                )
                context["ledger_data"] = ledger_data
                if request.GET.get("export") == "excel":
                    return self.export_general_ledger_excel(ledger_data)
            except Exception as e:
                # Handle errors (like invalid account code)
                from django.contrib import messages

                messages.error(request, f"Error: {str(e)}")

        return TemplateResponse(request, "admin/reports/general_ledger.html", context)
